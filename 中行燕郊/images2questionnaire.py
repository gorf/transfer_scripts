#!/usr/bin/python
# -*- coding: UTF-8 -*-
'''images2questionnaire.py
@version: $Id$
@author: U{Gorf Liu <liuqing.com@gmail.com>}
@license: GPL
@see: 参考资料链接等等
通过截图文件生成分网点问卷
'''

import sys
import os,string,pandas

import configparser
global error_number
error_number = 0
if os.path.isfile('扣分汇总.xlsx'):
    os.remove('扣分汇总.xlsx')

def read_config(config_file_path):
    cf = configparser.ConfigParser()
    cf.read(config_file_path, encoding='UTF-8')

    s = cf.sections()

    o = cf.options("baseconf")

    v = cf.items("baseconf")

    model_file = cf.get("baseconf", "model_file")
    model_fileA = cf.get("baseconf", "model_fileA")
    model_fileB = cf.get("baseconf", "model_fileB")
    enableq_file = cf.get("baseconf", "enableq_file")
    index_file = cf.get("baseconf", "index_file")
    all_amount_file = cf.get("baseconf", "all_amount_file")
    image_dir = cf.get("baseconf","image_dir")
    bank_type = cf.get("baseconf","bank_type")
    bank_spot_type = cf.get("baseconf","bank_spot_type")
    spot_names_file = cf.get("baseconf","spot_names")
    head_dir = cf.get("baseconf","head_dir")
    return(model_file,model_fileA,model_fileB,enableq_file,index_file,all_amount_file,image_dir,bank_type,bank_spot_type,
            spot_names_file, head_dir)
    
def browse_images(path):    
    print(path)
    if not os.path.isdir(path):
            print("错误：没有这个网点的目录，没扣分？太强了！")
            return

    for root, dirs, list in os.walk(path):
		#root遍历路径，dirs当前遍历路径下的目录，list当前遍历目录下的文件名
                for i in list:
                        #将分离的部分组成一个路径名
                        dir = os.path.join(root, i)
                        #print(i)
                return(list)

def main():
        '''主函数
            - 命令行测试
        @todo: 计划完成...
        '''
        #model_file = '附件4、富滇银行问卷-2017版-内部用-2017-03-09-1.xlsx'
        #try:
        #    enableq_file=sys.argv[1]
        #except:
        #    raise
        #    #enableq_file = 'Result_97_List_2016-09-08.csv'
        #index_file = 'Index_Result_Data_82_List_2016-06-08.csv'
        all_amount_file = '富滇银行2016年总表.xlsx'

        files = read_config("images2questionnaire.ini")
        #print(files)
        model_file = files[0].strip('\'')
        model_fileA = files[1].strip('\'')
        model_fileB = files[2].strip('\'')
        #print(model_file)
        enableq_file = files[3].strip('\'')
        index_file = files[4].strip('\'')
        all_amount_file = files[5].strip('\'')
        image_dir = files[6].strip('\'')
        bank_type = files[7].strip('\'')
        bank_spot_type = files[8].strip('\'')
        spot_names_file = files[9].strip('\'')
        head_dir = files[10].strip('\'')


        import csv
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font,Style
        from openpyxl.styles.colors import RED
        from openpyxl.drawing.image import Image
        import urllib.request


        if not os.path.exists('result'):
            os.makedirs('result')
        #if not os.path.exists('result_内部版'):
        #    os.makedirs('result_内部版')
        #if not os.path.exists('images'):
        #    os.makedirs('images')

        all_amount_wb = Workbook()
        all_amount_ws = all_amount_wb.active
        #all_amount_ws = all_amount_wb.create_sheet(title="总表")
        
        #写表头
        def read_head(spot_name):
            head = pandas.read_excel(os.path.join(head_dir,'中国银行-' + spot_name[4:] +
                        '-问卷.xls'))
            #print(head)
            date_survey = head.iloc[0,2].date()
            begin_end = head.iloc[0,5]
            business = head.iloc[1,0]
            manager = head.iloc[1,2]
            teller = head.iloc[1,5]
            return(date_survey,begin_end,business, manager, teller)
            
        #写图片数据的函数
        def write_images(spot_name,date_survey,begin_end,business, manager, teller):
            #print(index,spot_number,spot_name,point_type)

            #if point_type = '综合':
            #    model_file = model_fileA
            #elif point_type = '社区':
            #    model_file = model_fileB
            #else:
            #    print('网点类型错！')
            
            wb = openpyxl.load_workbook(model_file)
            ws = wb.active
            #写表头
            ws.cell(row=2,column=2).value = spot_name
            ws.cell(row=2,column=4).value = date_survey
            ws.cell(row=2,column=7).value = begin_end
            ws.cell(row=3,column=2).value = business
            ws.cell(row=3,column=4).value = manager
            ws.cell(row=3,column=7).value = teller

            global error_number
            questionnaire = pandas.read_excel(model_file).fillna(method='pad')
            #找图片文件，填三级指标分值
            errors = browse_images(os.path.join(os.getcwd(),image_dir, '中国银行-'+ spot_name[4:] + '-违规截图'))
            if os.path.isfile('扣分汇总.xlsx'):
                    all_errors = pandas.read_excel('扣分汇总.xlsx')
            else:
                    all_errors = pandas.DataFrame(columns =
                            ['网点名称','检查日期','进入离开网点时间',
                            '一级指标',	'二级指标', '序号', '细项指标',
                            '标准分值',	'扣分原因'])
                    #all_errors = pandas.DataFrame(columns = ['网点编号','网点名称','网点类型','检查日期','进入网点时间','离开网点时间','二级指标编号','二级指标描述','三级指标编号','三级指标描述','描述'])
            if errors:

                for error in errors:
                    error_number+=1
                    error_sp = error.split('-')
                    #print(error_sp)
                    standard_lev3 = error_sp[2]
                    error_des = '*'+error_sp[3].split('.')[0]
                    #print(standard_lev3,error_des)
                    row_number = -1
                    have_error = False
                    for row in ws.iter_rows():
                        #print(row[9].value)
                        if row[3].value == standard_lev3:
                            row[8].value = 0
                            all_errors.loc[-1] = [ spot_name,
                            date_survey,begin_end,questionnaire.iloc[row_number,0],questionnaire.iloc[row_number,1],standard_lev3,row[4].value, row[8].value, error_des]
                            all_errors.index=all_errors.index +1
                            if row[9].value:
                                row[9].value += '\n'+error_des
                            else:
                                row[9].value = error_des
                            have_error = True
                            break
                        row_number+=1
                    if not have_error:
                            all_errors.loc[-1] = [ spot_name,
                            date_survey,begin_end,'未找到匹配项','未找到匹配项','未找到匹配项',  standard_lev3, '未找到匹配项', error_des]
                            all_errors.index=all_errors.index +1
                            

            else: #没有扣分点
                            all_errors.loc[-1] = [ spot_name,
                            date_survey,begin_end,'这个网点没有找到，请检查','',
                            '', '', '', '']
                            all_errors.index=all_errors.index +1
            #print(all_errors)
            writer = pandas.ExcelWriter('扣分汇总.xlsx', engine='openpyxl')
            print('扣分条数',error_number)
            all_errors.to_excel(writer, startrow=0, index=False)
            writer.save()
            filename = os.path.join('result',spot_name + '-' +'问卷.xlsx')
            wb.save(filename)

            return(filename)
        #计算二级指标分值
        def cal_lev2(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            level2_value=lave3_value=0
            datang = 1
            if ws.cell(row=31,column=10).value:#大堂服务人员不在岗问题
                for row in range(32, 40):
                        ws.cell(row=row,column=9).value = 0

            wb.save(filename)
            #ws.column_dimensions.group('I','J', hidden=True)
            #ws.column_dimensions.group('N','O', hidden=True)
            #print(filename[11:])
            #wb.save(os.path.join('result',filename[11:]))

            #print(ws.title)

        
        ##读取enableQ文件
        #with open(enableq_file,'r', encoding='gb18030') as f:
        #    f_csv = csv.DictReader(f)
        #    header = f_csv.fieldnames
        #    ##headings = next(f_csv)
        #    ##Row = namedtuple('Row',headings)
        #    try:
        #        for row in f_csv:
        #            ##row = Row(*r)
        #            print(row['序号'])
        #            print(row['样本标识'])
        spot_names = pandas.read_excel(spot_names_file)        
        for index,row in spot_names.iterrows():
            date_survey,begin_end,business, manager, teller = read_head(row['网点名称'])
            filename = write_images(row['网点名称'],date_survey,begin_end,business,
                    manager, teller)
            cal_lev2(filename)
            #except UnicodeDecodeError:
            #    print(row['序号']+' error')
            #all_amount_wb.save(filename = all_amount_file)    
        #读取图片文件
        #bank_spots = pandas.read_excel(bank_spot_type)
        #for index,row in bank_spots.iterrows():
                
if __name__=="__main__":
        main()



